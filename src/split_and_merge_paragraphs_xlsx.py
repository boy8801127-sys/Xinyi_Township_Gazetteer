# -*- coding: utf-8 -*-
"""
段落 XLSX 處理（四步驟）：
0. 合併 A 欄孤兒空白 ID 與上一列
1. 依 P{數字} 分篇
2. 對每個 A 欄區塊，若段落／頁數跨列則合併 B、D 欄
3. 輸出各篇 xlsx 與合併總檔
"""
from __future__ import annotations

import re
import sys
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

try:
    from . import config
    from .merge_paragraph_rows import FIELDNAMES, _merge_page_ranges
except ImportError:
    config = None
    from merge_paragraph_rows import FIELDNAMES, _merge_page_ranges

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_DIR = ROOT / "output" if config is None else config.OUTPUT_DIR

_PAPER_ID_RE = re.compile(r"^P(\d+)-", re.IGNORECASE)

COL_ID = 1
COL_PARA = 2
COL_SOURCE = 3
COL_PAGES = 4

PAPER_FIELDNAMES = ["ID", "段落", "頁數"]


@dataclass
class PaperBlock:
    paper_index: str
    source_name: str
    row_indices: list[int] = field(default_factory=list)


@dataclass
class ProcessReport:
    a_merges_added: int = 0
    bd_merges_added: int = 0
    paper_count: int = 0


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_page(value: Any) -> str:
    s = _cell_str(value)
    if not s:
        return ""
    return s.replace("-", "–")


def _get_header_map(ws: Worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        h = _cell_str(ws.cell(1, col).value)
        if h in FIELDNAMES:
            header_map[h] = col
    missing = [f for f in FIELDNAMES if f not in header_map]
    if missing:
        raise ValueError(f"缺少必要欄位：{missing}")
    return header_map


def _row_has_content(ws: Worksheet, row: int, header_map: dict[str, int]) -> bool:
    for name in FIELDNAMES:
        col = header_map[name]
        val = ws.cell(row, col).value
        if name == "頁數":
            if _normalize_page(val):
                return True
        elif _cell_str(val):
            return True
    return False


def _merged_range_for_cell(ws: Worksheet, row: int, col: int) -> CellRange | None:
    for merged in ws.merged_cells.ranges:
        if (
            merged.min_col <= col <= merged.max_col
            and merged.min_row <= row <= merged.max_row
        ):
            return merged
    return None


def _id_cell_value(ws: Worksheet, row: int, col_id: int) -> str:
    """取得該列 ID；若落在 A 欄合併格內則取合併區塊左上角。"""
    merged = _merged_range_for_cell(ws, row, col_id)
    if merged:
        return _cell_str(ws.cell(merged.min_row, col_id).value)
    return _cell_str(ws.cell(row, col_id).value)


def _is_excel_merged_blank(ws: Worksheet, row: int, col_id: int) -> bool:
    merged = _merged_range_for_cell(ws, row, col_id)
    if not merged:
        return False
    if merged.min_col != col_id or merged.max_col != col_id:
        return False
    if row <= merged.min_row:
        return False
    return not _cell_str(ws.cell(row, col_id).value)


def _is_orphan_blank_id(ws: Worksheet, row: int, col_id: int) -> bool:
    if _cell_str(ws.cell(row, col_id).value):
        return False
    return not _is_excel_merged_blank(ws, row, col_id)


def _paper_index_from_id(row_id: str) -> str | None:
    m = _PAPER_ID_RE.match((row_id or "").strip())
    return m.group(1) if m else None


def merge_orphan_id_cells(ws: Worksheet, header_map: dict[str, int]) -> int:
    """Step 0：將孤兒空白 ID 列與上一列合併 A 欄。回傳新增 merge 數。"""
    col_id = header_map["ID"]
    merges_added = 0
    block_start: int | None = None
    pending_orphans: list[int] = []

    def flush_block() -> None:
        nonlocal merges_added, block_start, pending_orphans
        if block_start is not None and pending_orphans:
            end_row = pending_orphans[-1]
            if end_row > block_start:
                ws.merge_cells(
                    start_row=block_start,
                    start_column=col_id,
                    end_row=end_row,
                    end_column=col_id,
                )
                merges_added += 1
        block_start = None
        pending_orphans = []

    for row in range(2, ws.max_row + 1):
        if not _row_has_content(ws, row, header_map):
            continue

        id_val = _id_cell_value(ws, row, col_id)

        if id_val:
            flush_block()
            block_start = row
            pending_orphans = []
        elif _is_orphan_blank_id(ws, row, col_id):
            if block_start is not None:
                pending_orphans.append(row)
        # Excel 已合併空白：A 欄已處理，不加入 pending_orphans

    flush_block()
    return merges_added


def split_worksheet_by_paper(ws: Worksheet, header_map: dict[str, int]) -> dict[str, PaperBlock]:
    """Step 1：依 P{數字} 分篇。"""
    col_id = header_map["ID"]
    col_source = header_map["來源文章"]
    by_paper: dict[str, PaperBlock] = {}
    current_paper: str | None = None

    for row in range(2, ws.max_row + 1):
        if not _row_has_content(ws, row, header_map):
            continue

        id_val = _id_cell_value(ws, row, col_id)
        paper = _paper_index_from_id(id_val) if id_val else current_paper
        if paper is None:
            continue

        current_paper = paper
        if paper not in by_paper:
            by_paper[paper] = PaperBlock(paper_index=paper, source_name="")
        block = by_paper[paper]
        block.row_indices.append(row)

        source = _cell_str(ws.cell(row, col_source).value)
        if source and not block.source_name:
            block.source_name = source

    return by_paper


def _iter_a_blocks_in_rows(
    ws: Worksheet, rows: list[int], col_id: int
) -> list[tuple[int, int]]:
    """列出指定列範圍內的 A 欄區塊 (start, end)。"""
    if not rows:
        return []

    row_set = set(rows)
    visited: set[int] = set()
    blocks: list[tuple[int, int]] = []

    for row in sorted(rows):
        if row in visited:
            continue

        merged = _merged_range_for_cell(ws, row, col_id)
        if merged and merged.min_col == col_id and merged.max_col == col_id:
            start = merged.min_row
            end = merged.max_row
            block_rows = [r for r in range(start, end + 1) if r in row_set]
            if not block_rows:
                continue
            start, end = min(block_rows), max(block_rows)
        else:
            start = end = row

        for r in range(start, end + 1):
            visited.add(r)
        blocks.append((start, end))

    return blocks


def _column_already_merged(ws: Worksheet, row_start: int, row_end: int, col: int) -> bool:
    if row_end <= row_start:
        return False
    for merged in ws.merged_cells.ranges:
        if (
            merged.min_col == col
            and merged.max_col == col
            and merged.min_row == row_start
            and merged.max_row == row_end
        ):
            return True
    return False


def _unmerge_column_range(
    ws: Worksheet, row_start: int, row_end: int, col: int
) -> None:
    """解除指定列範圍內某欄的垂直合併。"""
    to_remove: list[str] = []
    for merged in ws.merged_cells.ranges:
        if merged.min_col != col or merged.max_col != col:
            continue
        if merged.max_row < row_start or merged.min_row > row_end:
            continue
        to_remove.append(str(merged))
    for ref in to_remove:
        ws.unmerge_cells(ref)


def _set_cell_value(ws: Worksheet, row: int, col: int, value: Any) -> None:
    """寫入儲存格；若為合併格非左上角，先解除該格所在合併。"""
    merged = _merged_range_for_cell(ws, row, col)
    if merged and (row != merged.min_row or col != merged.min_col):
        ws.unmerge_cells(str(merged))
    ws.cell(row, col).value = value


def merge_block_paragraph_and_pages(
    ws: Worksheet,
    row_start: int,
    row_end: int,
    header_map: dict[str, int],
) -> int:
    """Step 2：若段落／頁數跨列，合併 B、D 欄。回傳新增 merge 數。"""
    if row_end <= row_start:
        return 0

    col_para = header_map["段落"]
    col_pages = header_map["頁數"]
    merges_added = 0

    has_spread = False
    for row in range(row_start + 1, row_end + 1):
        if _cell_str(ws.cell(row, col_para).value) or _normalize_page(
            ws.cell(row, col_pages).value
        ):
            has_spread = True
            break

    if not has_spread:
        return 0

    paragraphs: list[str] = []
    page_strs: list[str] = []
    for row in range(row_start, row_end + 1):
        para = _cell_str(ws.cell(row, col_para).value)
        pages = _normalize_page(ws.cell(row, col_pages).value)
        if para:
            paragraphs.append(para)
        if pages:
            page_strs.append(pages)

    merged_para = "。".join(paragraphs)
    merged_pages = _merge_page_ranges(page_strs)

    _unmerge_column_range(ws, row_start, row_end, col_para)
    _unmerge_column_range(ws, row_start, row_end, col_pages)

    _set_cell_value(ws, row_start, col_para, merged_para)
    _set_cell_value(ws, row_start, col_pages, merged_pages)
    for row in range(row_start + 1, row_end + 1):
        _set_cell_value(ws, row, col_para, None)
        _set_cell_value(ws, row, col_pages, None)

    if not _column_already_merged(ws, row_start, row_end, col_para):
        ws.merge_cells(
            start_row=row_start,
            start_column=col_para,
            end_row=row_end,
            end_column=col_para,
        )
        merges_added += 1

    if not _column_already_merged(ws, row_start, row_end, col_pages):
        ws.merge_cells(
            start_row=row_start,
            start_column=col_pages,
            end_row=row_end,
            end_column=col_pages,
        )
        merges_added += 1

    return merges_added


def process_paper_blocks(
    ws: Worksheet, header_map: dict[str, int], by_paper: dict[str, PaperBlock]
) -> int:
    """對每篇執行 Step 2（B/D 欄合併）。來源文章維持逐列，不合併。"""
    col_id = header_map["ID"]
    bd_merges = 0

    for paper in sorted(by_paper.keys(), key=lambda x: int(x)):
        block = by_paper[paper]
        rows = block.row_indices
        if not rows:
            continue

        for row_start, row_end in _iter_a_blocks_in_rows(ws, rows, col_id):
            bd_merges += merge_block_paragraph_and_pages(
                ws, row_start, row_end, header_map
            )

    return bd_merges


def _effective_cell_value(ws: Worksheet, row: int, col: int) -> Any:
    """讀取儲存格值；若為合併格非左上角，取合併區塊左上角。"""
    merged = _merged_range_for_cell(ws, row, col)
    if merged:
        return ws.cell(merged.min_row, merged.min_col).value
    return ws.cell(row, col).value


def _copy_row_range_with_merges(
    src_ws: Worksheet,
    dst_ws: Worksheet,
    src_rows: list[int],
    dst_start_row: int,
    max_col: int = 4,
) -> None:
    """複製列與合併儲存格至目標工作表。"""
    if not src_rows:
        return

    row_offset = dst_start_row - min(src_rows)
    src_row_set = set(src_rows)

    for src_row in src_rows:
        dst_row = src_row + row_offset
        for col in range(1, max_col + 1):
            src_cell = src_ws.cell(src_row, col)
            dst_cell = dst_ws.cell(dst_row, col)
            dst_cell.value = _effective_cell_value(src_ws, src_row, col)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = copy(src_cell.number_format)
                dst_cell.protection = copy(src_cell.protection)
                dst_cell.alignment = copy(src_cell.alignment)

    for merged in src_ws.merged_cells.ranges:
        if merged.max_col > max_col:
            continue
        # 來源文章（C 欄）維持逐列，不複製垂直合併
        if merged.min_col == COL_SOURCE and merged.max_col == COL_SOURCE:
            continue
        if merged.min_row not in src_row_set or merged.max_row not in src_row_set:
            continue
        dst_ws.merge_cells(
            start_row=merged.min_row + row_offset,
            start_column=merged.min_col,
            end_row=merged.max_row + row_offset,
            end_column=merged.max_col,
        )


def _write_header(dst_ws: Worksheet) -> None:
    for col, name in enumerate(FIELDNAMES, start=1):
        dst_ws.cell(1, col, name)


def _write_paper_header(dst_ws: Worksheet) -> None:
    """個別論文檔：僅 ID、段落、頁數（無來源文章）。"""
    for col, name in enumerate(PAPER_FIELDNAMES, start=1):
        dst_ws.cell(1, col, name)


def _copy_paper_rows_without_source(
    src_ws: Worksheet,
    dst_ws: Worksheet,
    src_rows: list[int],
    dst_start_row: int,
) -> None:
    """複製至個別論文檔：刪除來源文章欄，頁數原封不動移至 C 欄。"""
    if not src_rows:
        return

    row_offset = dst_start_row - min(src_rows)
    src_row_set = set(src_rows)

    col_map = {COL_ID: 1, COL_PARA: 2, COL_PAGES: 3}

    for src_row in src_rows:
        dst_row = src_row + row_offset
        for src_col, dst_col in col_map.items():
            src_cell = src_ws.cell(src_row, src_col)
            dst_cell = dst_ws.cell(dst_row, dst_col)
            dst_cell.value = _effective_cell_value(src_ws, src_row, src_col)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = copy(src_cell.number_format)
                dst_cell.protection = copy(src_cell.protection)
                dst_cell.alignment = copy(src_cell.alignment)

    for merged in src_ws.merged_cells.ranges:
        if merged.min_row not in src_row_set or merged.max_row not in src_row_set:
            continue
        if merged.min_col == COL_SOURCE and merged.max_col == COL_SOURCE:
            continue
        if merged.min_col == COL_ID and merged.max_col == COL_ID:
            dst_col = 1
        elif merged.min_col == COL_PARA and merged.max_col == COL_PARA:
            dst_col = 2
        elif merged.min_col == COL_PAGES and merged.max_col == COL_PAGES:
            dst_col = 3
        else:
            continue
        dst_ws.merge_cells(
            start_row=merged.min_row + row_offset,
            start_column=dst_col,
            end_row=merged.max_row + row_offset,
            end_column=dst_col,
        )


def _sanitize_filename(name: str, max_len: int = 80) -> str:
    s = re.sub(r'[<>:"/\\|?*]', "_", name)
    s = s.strip(" .") or "未命名"
    return s[:max_len]


def process_paragraphs_xlsx(
    input_path: Path,
    output_dir: Path | None = None,
    date_suffix: str | None = None,
) -> tuple[Path, list[Path], ProcessReport]:
    """
    執行完整四步驟流程。

    回傳：(合併總檔路徑, 各篇路徑列表, 處理報告)
    """
    input_path = Path(input_path)
    base_dir = output_dir or input_path.parent
    by_paper_dir = base_dir / "paragraphs_by_paper_xlsx"
    by_paper_dir.mkdir(parents=True, exist_ok=True)

    if date_suffix is None:
        stem = input_path.stem
        if stem.startswith("paragraphs_all_"):
            date_suffix = stem.replace("paragraphs_all_", "")
        else:
            date_suffix = stem

    wb = load_workbook(input_path, data_only=False)
    ws = wb.active
    header_map = _get_header_map(ws)
    report = ProcessReport()

    report.a_merges_added = merge_orphan_id_cells(ws, header_map)
    by_paper = split_worksheet_by_paper(ws, header_map)
    report.paper_count = len(by_paper)
    report.bd_merges_added = process_paper_blocks(ws, header_map, by_paper)

    all_path = base_dir / f"paragraphs_all_merged_{date_suffix}.xlsx"
    combined_wb = Workbook()
    combined_ws = combined_wb.active
    combined_ws.title = "paragraphs"
    _write_header(combined_ws)
    next_row = 2

    paper_paths: list[Path] = []
    for paper_idx in sorted(by_paper.keys(), key=lambda x: int(x)):
        block = by_paper[paper_idx]
        rows = block.row_indices
        if not rows:
            continue

        safe = _sanitize_filename(block.source_name) if block.source_name else f"paper_{paper_idx}"
        out_name = f"P{paper_idx}_{safe}_{date_suffix}.xlsx"
        out_path = by_paper_dir / out_name

        paper_wb = Workbook()
        paper_ws = paper_wb.active
        paper_ws.title = "paragraphs"
        _write_paper_header(paper_ws)
        _copy_paper_rows_without_source(ws, paper_ws, rows, dst_start_row=2)
        paper_wb.save(out_path)
        paper_wb.close()
        paper_paths.append(out_path)

        _copy_row_range_with_merges(ws, combined_ws, rows, dst_start_row=next_row)
        next_row += len(rows)

    combined_wb.save(all_path)
    combined_wb.close()
    wb.close()

    return all_path, paper_paths, report


def main() -> int:
    if len(sys.argv) < 2:
        input_path = DEFAULT_OUTPUT_DIR / "paragraphs_all_2026_06_17.xlsx"
    else:
        input_path = Path(sys.argv[1])

    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else None

    if not input_path.exists():
        print(f"檔案不存在: {input_path}", file=sys.stderr)
        return 1

    all_path, paper_paths, report = process_paragraphs_xlsx(input_path, output_dir)
    print(f"已輸出合併總檔: {all_path}")
    print(f"  論文篇數: {report.paper_count}")
    print(f"  Step 0 新增 A 欄合併: {report.a_merges_added}")
    print(f"  Step 2 新增 B/D 欄合併: {report.bd_merges_added}")
    print(f"  分檔 ({len(paper_paths)} 個):")
    for p in paper_paths:
        print(f"    - {p.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
